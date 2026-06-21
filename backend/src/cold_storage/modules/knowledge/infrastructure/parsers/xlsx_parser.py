"""XLSX file parser — openpyxl read-only mode with row/sheet limits."""

from __future__ import annotations

import io
import unicodedata

from cold_storage.modules.knowledge.domain.models import ParsedBlock
from cold_storage.modules.knowledge.infrastructure.parsers.base import (
    PARSER_VERSION,
    ParseResult,
    register_parser,
)

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None

MAX_XLSX_ROWS_PER_SHEET: int = 5_000
MAX_XLSX_SHEETS: int = 50
MAX_XLSX_COLUMNS: int = 200


class XlsxParser:
    """Parse .xlsx files using openpyxl in read-only mode.

    Records sheet_name and row ranges for each block.
    Enforces limits on rows, sheets, and columns.
    """

    name: str = "xlsx"

    def parse(self, content: bytes, filename: str) -> ParseResult:
        """Parse an .xlsx file into structured blocks.

        Raises
        ------
        ImportError
            If openpyxl is not installed.
        """
        if load_workbook is None:
            raise ImportError("openpyxl is required for XLSX parsing: pip install openpyxl")

        buf = io.BytesIO(content)
        wb = load_workbook(buf, read_only=True, data_only=False)

        blocks: list[ParsedBlock] = []
        order = 0
        sheet_count = 0

        try:
            for sheet_name in wb.sheetnames:
                sheet_count += 1
                if sheet_count > MAX_XLSX_SHEETS:
                    break

                ws = wb[sheet_name]
                row_num = 0
                headers: list[str] = []

                for row in ws.iter_rows(max_row=MAX_XLSX_ROWS_PER_SHEET + 1):
                    row_num += 1
                    if row_num > MAX_XLSX_ROWS_PER_SHEET:
                        break

                    # Trim columns
                    cells_raw = [cell.value for cell in row[:MAX_XLSX_COLUMNS]]
                    # Convert all to strings
                    cells = [str(c).strip() if c is not None else "" for c in cells_raw]

                    # Skip entirely empty rows
                    if not any(cells):
                        continue

                    # First row is treated as headers
                    if row_num == 1:
                        headers = cells
                        header_text = " | ".join(c for c in headers if c)
                        if header_text.strip():
                            blocks.append(
                                ParsedBlock(
                                    text=unicodedata.normalize("NFKC", header_text),
                                    block_type="metadata",
                                    section_path=f"sheet:{sheet_name}",
                                    page_start=None,
                                    page_end=None,
                                    sheet_name=sheet_name,
                                    row_start=1,
                                    row_end=1,
                                    source_order=order,
                                    metadata={
                                        "headers": headers,
                                        "parser_version": PARSER_VERSION,
                                    },
                                )
                            )
                            order += 1
                        continue

                    # Data row — build structured text with header mapping
                    pairs: list[str] = []
                    for i, cell_val in enumerate(cells):
                        if i < len(headers) and headers[i]:
                            pairs.append(f"{headers[i]}: {cell_val}")
                        elif cell_val:
                            pairs.append(cell_val)

                    row_text = " | ".join(p for p in pairs if p)
                    if not row_text.strip():
                        continue

                    row_text = unicodedata.normalize("NFKC", row_text)
                    blocks.append(
                        ParsedBlock(
                            text=row_text,
                            block_type="paragraph",
                            section_path=f"sheet:{sheet_name}",
                            page_start=None,
                            page_end=None,
                            sheet_name=sheet_name,
                            row_start=row_num,
                            row_end=row_num,
                            source_order=order,
                            metadata={
                                "parser_version": PARSER_VERSION,
                                "row_number": row_num,
                                "sheet_name": sheet_name,
                            },
                        )
                    )
                    order += 1
        finally:
            wb.close()

        return ParseResult(blocks=blocks)


register_parser(".xlsx", XlsxParser())
